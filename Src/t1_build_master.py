"""
t1_build_master.py  (TASK T1)
Parse the four sources, recompute maturity on ONE convention, assemble master dataset.

Outputs (in _artifacts/):
  lit119.csv        full 'Collected' sheet (isothermal literature, ambient)
  lit55.csv         the paper's exact 55-point baseline subset (reproduction)
  lit503.csv        large literature DB (91 mix families, 32 references) -> L3 pool
  slab.csv          in-situ slab (Table 3): SC / AD / Cored, core equivalent age
  cube.csv          400mm cube age-strength matrix (bridge; std-cured vs core)
  master.parquet    all sources unified with M_NS, t_eq, DD, physics features
  ea_calibration.csv  per-mix ASTM C1074 apparent activation energy

Run:  uv run python t1_build_master.py
"""
import os
import numpy as np
import pandas as pd
import openpyxl

from config import XLSX, OUT_DIR, CSV_ENC
import maturity as mat

pd.set_option("display.width", 200)
pd.set_option("display.max_columns", 40)


# ==========================================================================
# 1. 'Collected' sheet (119 rows) + identify paper's 55-point subset
# ==========================================================================
def load_collected():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Collected"]
    cols = ["time_h", "T_int", "T_amb", "RH", "WB", "Sa", "WC", "W", "C",
            "GGBFS", "FlyAsh", "fine", "coarse", "curing_C", "SP", "f_c",
            "slump", "air", "DD_orig"]
    rows = []
    for r in range(2, ws.max_row + 1):
        v = [ws.cell(r, c).value for c in range(1, 20)]
        if v[0] is None:
            continue
        rows.append(v)
    df = pd.DataFrame(rows, columns=cols).apply(pd.to_numeric, errors="coerce")

    # --- repair obvious time typos using the internally-consistent DD column ---
    # DD_orig == curing_C * time_h / 24 (verified). Where |recomputed-DD| is large,
    # the time cell is corrupt (e.g. row 'F@5C' shows 35556.32 vs true 3556.32).
    dd_check = df["curing_C"] * df["time_h"] / 24.0
    bad = (df["DD_orig"] > 0) & ((dd_check - df["DD_orig"]).abs() > 0.5 * df["DD_orig"])
    n_fix = int(bad.sum())
    df.loc[bad, "time_h"] = df.loc[bad, "DD_orig"] * 24.0 / df.loc[bad, "curing_C"]

    # binder = cement + scm ; mix signature
    df["GGBFS"] = df["GGBFS"].fillna(0)
    df["FlyAsh"] = df["FlyAsh"].fillna(0)
    df["binder"] = df["C"] + df["GGBFS"] + df["FlyAsh"]
    df["mix_id"] = ("LIT_wc" + (df["WC"] * 100).round().astype(int).astype(str)
                    + "_C" + df["C"].round().astype(int).astype(str)
                    + "_G" + df["GGBFS"].round().astype(int).astype(str)
                    + "_F" + df["FlyAsh"].round().astype(int).astype(str))
    df["age_day"] = df["time_h"] / 24.0
    return df, n_fix


def flag_paper55(df):
    """Paper's 55: mixes {A:wc.50/C330, B:.45/C370, C:.41/C410} at curing {5,20,40}
       plus {F:.55/C349/FA62, G:.55/C287/FA123} at curing 5 only. (GGBFS mixes excluded.)"""
    A = (df.C.round() == 330) & (df.GGBFS == 0) & (df.FlyAsh == 0)
    B = (df.C.round() == 370) & (df.GGBFS == 0) & (df.FlyAsh == 0)
    C = (df.C.round() == 410) & (df.GGBFS == 0) & (df.FlyAsh == 0)
    F = (df.C.round() == 349) & (df.FlyAsh.round() == 62)
    G = (df.C.round() == 287) & (df.FlyAsh.round() == 123)
    abc = (A | B | C) & df.curing_C.isin([5, 20, 40])
    fg = (F | G) & (df.curing_C == 5)
    return abc | fg


# ==========================================================================
# 2. Large literature DB (sheet index 1: 503 rows)  -> L3 pool
# ==========================================================================
def load_lit503():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.worksheets[1]
    hdr = [str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    data = []
    for r in range(2, ws.max_row + 1):
        v = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(x is not None and str(x).strip() for x in v):
            data.append(v)
    df = pd.DataFrame(data, columns=[h.strip() for h in hdr])
    df = df.rename(columns={
        "MIX": "mix_id", "Time (t) (days)": "age_day", "T_int (°C)": "T_int",
        "T_amb (°C)": "T_amb", "RH (%)": "RH", "W/B": "WB", "S/a (%)": "Sa",
        "w/c": "WC", "W (kg/m³)": "W", "Cement (kg/m³)": "C", "FA (kg/m³)": "FlyAsh",
        "SCM (kg/m³)": "SCM", "Fine Agg (kg/m³)": "fine", "Coarse Agg (kg/m³)": "coarse",
        "Curing (°C)": "curing_C", "Admix (SP, % bwoc)": "SP", "Strength (MPa)": "f_c",
        "Slump (mm)": "slump", "Air content (%)": "air", "Maturity (°D·D)": "DD_orig",
        "Reference": "ref"})
    for c in ["age_day", "T_int", "T_amb", "RH", "WB", "Sa", "WC", "W", "C",
              "FlyAsh", "SCM", "fine", "coarse", "curing_C", "SP", "f_c",
              "slump", "air", "DD_orig"]:
        if c in df:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    # In the source sheet the MIX name (and mix-design values) are written ONCE on the
    # first row of each mix-series and left blank for its subsequent age rows. Forward-
    # fill so every age row inherits its mix identity & design -> correct grouping.
    df["mix_id"] = df["mix_id"].astype("object").where(df["mix_id"].notna()).ffill()
    df["ref"] = df["ref"].astype("object").where(df["ref"].notna()).ffill()
    design_cols = ["T_int", "T_amb", "RH", "WB", "Sa", "WC", "W", "C", "FlyAsh",
                   "SCM", "fine", "coarse", "curing_C", "SP", "slump", "air"]
    df[design_cols] = df.groupby("mix_id")[design_cols].ffill()
    df = df[df["mix_id"].notna()].reset_index(drop=True)
    df["mix_family"] = df["mix_id"].astype(str).str.replace(r"[-–]\s*\d+.*$", "", regex=True).str.strip()
    df["time_h"] = df["age_day"] * 24.0
    df["FlyAsh"] = df.get("FlyAsh", 0).fillna(0)
    df["SCM"] = df.get("SCM", 0).fillna(0)
    df["binder"] = df["C"].fillna(0) + df["FlyAsh"] + df["SCM"]
    return df


# ==========================================================================
# 3. Slab in-situ (Table 3, verified from PDF) - the out-of-domain test set
#    curing_type: SC=standard, AD=air-dried, Cored=in-situ core (temp_source=core)
#    te = Arrhenius equivalent age (hours) from FBG core temperature (paper).
# ==========================================================================
SLAB = [
    # exp, age_day, te_h, SC, AD, Cored
    ("Exp1", 3,  123.83, 13.89, 13.77, 13.00),
    ("Exp1", 7,  266.22, 17.11, 16.08, 15.90),
    ("Exp1", 28, 947.59, 24.12, 22.41, 20.90),
    ("Exp2", 1,   28.86, np.nan, 10.94, np.nan),
    ("Exp2", 3,   80.19, 23.67, 22.07, 21.30),
    ("Exp2", 7,  163.63, 27.91, 26.85, 27.20),
    ("Exp2", 14, 310.15, 35.25, 28.81, 28.73),
    ("Exp2", 28, 572.11, 36.24, 31.21, 33.00),
]
SLAB_META = {  # from paper: mix design & average core temperature
    "Exp1": dict(WC=0.64, fck=24, avgT=27.79, RH=0.515, C=None),
    "Exp2": dict(WC=0.44, fck=32, avgT=16.47, RH=0.515, C=None),
}


def load_slab():
    recs = []
    for exp, ad, te, sc, adv, cored in SLAB:
        for ctype, val, tsrc in [("SC", sc, "standard"), ("AD", adv, "air"),
                                 ("Cored", cored, "core")]:
            if val is None or (isinstance(val, float) and np.isnan(val)):
                continue
            m = SLAB_META[exp]
            recs.append(dict(dataset="slab", mix_id=f"slab_{exp}", mix_family=f"slab_{exp}",
                             ref="Park et al. 2026", temp_source=tsrc, curing_type=ctype,
                             age_day=ad, time_h=ad * 24.0, t_eq=te, curing_C=m["avgT"],
                             WC=m["WC"], fck=m["fck"], RH=m["RH"] * 100, C=np.nan,
                             binder=np.nan, f_c=val))
    return pd.DataFrame(recs)


# ==========================================================================
# 4. 400mm cube (bridge). Values read from cube report + PNG (cross-checked).
#    std-cured (air '기건' / water '수중')  vs  cores (indoor '실내' / outdoor '실외').
#    A: nominal w/c 0.44 -> EFFECTIVE 0.555 (report re-analysis, user-confirmed).
# ==========================================================================
CUBE = {
    # eff_wc : { fck, air{age:MPa}, water{age:MPa}, core_indoor_28, core_outdoor_28 }
    0.555: dict(fck=32,
                air={1: 11.7, 3: 22.0, 7: 26.6, 14: 28.5, 28: 26.0},
                water={3: 24.8, 7: 28.8, 14: 31.3, 28: 27.2},
                core_indoor=25.8, core_outdoor=28.3, peakT=41.2, peak_day=0.72),
    # NOTE (2026-07-26): the 3-day pair was corrected against the strength-test figure
    # '2026_0611_압축강도_시험결과_rik.png' (labelled 이상치 제거 후 = after outlier removal),
    # which reports mean +- CoV per bar. It gives air 18.9 +-1.9% and water 19.3 +-6.0%; the
    # previous entries (air 19.3, water 20.4) had the water value in the air slot and a water
    # value that appears nowhere in the figure. All other 31 cube strengths match the figure.
    0.50: dict(fck=24,
               air={1: 8.1, 3: 18.9, 7: 20.7, 14: 21.7, 28: 23.5},
               water={3: 19.3, 7: 22.4, 14: 23.4, 28: 30.2},
               core_indoor=29.9, core_outdoor=24.8, peakT=41.2, peak_day=0.83),
    0.60: dict(fck=18,
               air={1: 3.5, 3: 14.4, 7: 13.9, 14: 16.1, 28: 14.9},
               water={3: 14.4, 7: 16.6, 14: 16.3, 28: 19.5},
               core_indoor=22.9, core_outdoor=24.7, peakT=37.9, peak_day=0.92),
}
CUBE_AMBIENT_C = 24.0    # provisional standard-curing temperature (May outdoor avg)


def load_cube():
    recs = []
    for wc, d in CUBE.items():
        mid = f"cube_wc{int(wc*1000)}"
        for age, v in d["air"].items():
            recs.append(dict(dataset="cube", mix_id=mid, mix_family=mid, ref="This study (cube)",
                             temp_source="standard", curing_type="air", age_day=age,
                             time_h=age*24, curing_C=CUBE_AMBIENT_C, WC=wc, fck=d["fck"],
                             peakT=d["peakT"], peak_day=d["peak_day"], f_c=v))
        for age, v in d["water"].items():
            recs.append(dict(dataset="cube", mix_id=mid, mix_family=mid, ref="This study (cube)",
                             temp_source="standard", curing_type="water", age_day=age,
                             time_h=age*24, curing_C=20.0, WC=wc, fck=d["fck"],
                             peakT=d["peakT"], peak_day=d["peak_day"], f_c=v))
        for lab, v in [("core_indoor", d["core_indoor"]), ("core_outdoor", d["core_outdoor"])]:
            recs.append(dict(dataset="cube", mix_id=mid, mix_family=mid, ref="This study (cube)",
                             temp_source="core", curing_type=lab, age_day=28, time_h=28*24,
                             curing_C=d["peakT"], WC=wc, fck=d["fck"], peakT=d["peakT"],
                             peak_day=d["peak_day"], f_c=v))
    return pd.DataFrame(recs)


# ==========================================================================
# 5. Maturity recomputation + assembly
# ==========================================================================
def add_iso_maturity(df):
    """Isothermal literature rows: use curing temperature over the full age."""
    T = df["curing_C"].fillna(df.get("T_amb"))
    df["M_NS"] = mat.nurse_saul_iso(T, df["time_h"])
    df["t_eq"] = mat.eq_age_iso(T, df["time_h"])
    df["DD"] = mat.degree_day_iso(T, df["time_h"])
    df["provisional_maturity"] = False
    return df


def add_physics_features(df):
    df["sqrt_DD"] = np.sqrt(df["DD"].clip(lower=0))
    df["log_time"] = np.log(df["time_h"].clip(lower=1e-3))
    df["t_eq_day"] = df["t_eq"] / 24.0
    df["alpha_bazant"] = mat.hydration_degree_bazant(df["t_eq_day"])
    df["alpha_fh"] = mat.hydration_degree_fh(df["t_eq_day"])
    return df


def main():
    print("=" * 78)
    print("TASK T1 - Unified maturity recomputation & master dataset")
    print("=" * 78)

    # ---- literature: 119, 55, 503 ----
    coll, n_fix = load_collected()
    coll["dataset"] = "lit119"
    coll["temp_source"] = "ambient"
    coll["ref"] = "Ryu et al. 2024"
    coll["mix_family"] = coll["mix_id"]
    coll = add_iso_maturity(coll)
    print(f"\n[Collected sheet] rows={len(coll)}  time-typos repaired via DD consistency: {n_fix}")

    is55 = flag_paper55(coll)
    lit55 = coll[is55].copy()
    lit55["dataset"] = "lit55"
    print(f"[Paper 55-point subset] rows={len(lit55)}  (expected 55) -> "
          f"{'OK' if len(lit55)==55 else 'CHECK'};  unique mixes={lit55.mix_id.nunique()}")
    print("   mixes:", sorted(lit55.mix_id.unique()))
    # validate my DD recompute vs paper's DD_orig on the 55
    dd_err = (lit55["DD"] - lit55["DD_orig"]).abs().max()
    print(f"   max|DD_recomputed - DD_paper| on 55 = {dd_err:.3f}  (should be ~0)")

    lit503 = load_lit503()
    lit503["dataset"] = "lit503"
    lit503["temp_source"] = np.where(
        (lit503["T_int"] - lit503["T_amb"]).abs() > 0.5, "core", "ambient")
    lit503 = add_iso_maturity(lit503)
    print(f"\n[Large literature DB] rows={len(lit503)}  mix_families={lit503.mix_family.nunique()}"
          f"  refs={lit503.ref.nunique()}")
    print(f"   strength MPa: {lit503.f_c.min():.1f}-{lit503.f_c.max():.1f}; "
          f"w/c: {lit503.WC.min():.3f}-{lit503.WC.max():.3f}; "
          f"temp_source ambient/core = {(lit503.temp_source=='ambient').sum()}/"
          f"{(lit503.temp_source=='core').sum()}")

    # ---- slab & cube ----
    slab = load_slab()
    cube = load_cube()
    # provisional maturity for slab (te given from core; M_NS/DD from avg core temp)
    slab["M_NS"] = mat.nurse_saul_iso(slab["curing_C"], slab["time_h"])
    slab["DD"] = mat.degree_day_iso(slab["curing_C"], slab["time_h"])
    slab["provisional_maturity"] = slab["curing_type"] != "Cored"  # Cored te is exact (paper)
    slab.loc[slab.curing_type == "Cored", "provisional_maturity"] = False
    cube["t_eq"] = mat.eq_age_iso(cube["curing_C"], cube["time_h"])
    cube["M_NS"] = mat.nurse_saul_iso(cube["curing_C"], cube["time_h"])
    cube["DD"] = mat.degree_day_iso(cube["curing_C"], cube["time_h"])
    cube["provisional_maturity"] = True   # overwritten below where exact logs exist
    # ---- cube CORE rows: replace the constant-peak placeholder with the maturity
    #      integrated from the real 2-min thermocouple logs (t8_core_logs.py ->
    #      core_maturity.csv). The placeholder held the core at its peak temperature
    #      for the full 28 d, overstating t_eq by ~2.3x; the integrated history is
    #      the physically meaningful value.
    cm_path = os.path.join(OUT_DIR, "core_maturity.csv")
    if os.path.exists(cm_path):
        cm = pd.read_csv(cm_path, encoding="utf-8-sig")
        # indoor cores <- members 1-3 (member="cube"); outdoor core <- member 4 ("cube_out")
        n_fixed = 0
        for member, ctype in [("cube", "core_indoor"), ("cube_out", "core_outdoor")]:
            sub = cm[cm.member == member]
            if sub.empty and member == "cube_out":
                sub = cm[cm.member == "cube"]   # fallback: old single-history CSV
            for _, r in sub.iterrows():
                sel = ((cube.curing_type == ctype) & (cube.WC == r.wc)
                       & (cube.age_day == r.age_d))
                if sel.any():
                    cube.loc[sel, ["M_NS", "t_eq", "DD"]] = (
                        r.M_NS_core, r.teq_core, r.DD_core)
                    cube.loc[sel, "provisional_maturity"] = False
                    n_fixed += int(sel.sum())
        print(f"[400mm cube] core maturity from integrated logs for {n_fixed} rows "
              f"(indoor: TC 1-3 mean; outdoor: TC 4)")
    else:
        print("[400mm cube] WARNING: core_maturity.csv not found - core rows keep "
              "the constant-peak PLACEHOLDER maturity (run t8_core_logs.py first)")
    print(f"\n[Slab in-situ] rows={len(slab)}  (Cored={sum(slab.curing_type=='Cored')}, "
          f"SC={sum(slab.curing_type=='SC')}, AD={sum(slab.curing_type=='AD')})")
    print(f"[400mm cube] rows={len(cube)}  std-cured={sum(cube.temp_source=='standard')}, "
          f"cores={sum(cube.temp_source=='core')}")

    # ---- unify ----
    keep = ["dataset", "mix_id", "mix_family", "ref", "temp_source", "curing_type",
            "curing_C", "age_day", "time_h", "M_NS", "t_eq", "DD",
            "WC", "WB", "C", "binder", "RH", "Sa", "SP", "fck",
            "peakT", "peak_day", "provisional_maturity", "f_c"]
    frames = []
    for d in [lit55, coll, lit503, slab, cube]:
        for k in keep:
            if k not in d:
                d[k] = np.nan
        frames.append(d[keep])
    master = pd.concat(frames, ignore_index=True)
    master = master[master["f_c"].notna()].reset_index(drop=True)
    master = add_physics_features(master)

    # ---- Ea calibration (ASTM C1074) for mixes A,B,C (3 temps each) ----
    ea_rows = []
    for mid, g in coll.groupby("mix_id"):
        temps = sorted(g.curing_C.dropna().unique())
        if len(temps) >= 2:
            mrows = {t: (g[g.curing_C == t]["time_h"].values,
                         g[g.curing_C == t]["f_c"].values) for t in temps}
            res = mat.calibrate_Ea(mrows)
            ea_rows.append(dict(mix_id=mid, n_temp=res["n_temp"], Ea_Jmol=res["Ea"],
                                r2_arrhenius=res.get("r2_arrhenius", np.nan)))
    ea_df = pd.DataFrame(ea_rows)
    print("\n[ASTM C1074 apparent activation energy calibration]  (paper fixes Ea=38,300)")
    if len(ea_df):
        valid = ea_df[ea_df.Ea_Jmol.notna() & (ea_df.n_temp >= 3)]
        print(ea_df.to_string(index=False))
        if len(valid):
            print(f"   -> calibrated Ea (mixes with 3 temps): "
                  f"mean={valid.Ea_Jmol.mean():.0f} J/mol, "
                  f"range {valid.Ea_Jmol.min():.0f}-{valid.Ea_Jmol.max():.0f}; "
                  f"paper default=38,300")

    # ---- write ----
    lit55.to_csv(os.path.join(OUT_DIR, "lit55.csv"), index=False, encoding=CSV_ENC)
    coll.to_csv(os.path.join(OUT_DIR, "lit119.csv"), index=False, encoding=CSV_ENC)
    lit503.to_csv(os.path.join(OUT_DIR, "lit503.csv"), index=False, encoding=CSV_ENC)
    slab.to_csv(os.path.join(OUT_DIR, "slab.csv"), index=False, encoding=CSV_ENC)
    cube.to_csv(os.path.join(OUT_DIR, "cube.csv"), index=False, encoding=CSV_ENC)
    ea_df.to_csv(os.path.join(OUT_DIR, "ea_calibration.csv"), index=False, encoding=CSV_ENC)
    master.to_parquet(os.path.join(OUT_DIR, "master.parquet"), index=False)

    print("\n" + "=" * 78)
    print("MASTER DATASET SUMMARY")
    print("=" * 78)
    print(master.groupby(["dataset", "temp_source"]).agg(
        n=("f_c", "size"), mixes=("mix_id", "nunique"),
        fc_min=("f_c", "min"), fc_max=("f_c", "max")).to_string())
    print(f"\nTotal usable rows (f_c present): {len(master)}")
    print(f"Written -> {OUT_DIR}\\{{lit55,lit119,lit503,slab,cube,ea_calibration}}.csv, master.parquet")


if __name__ == "__main__":
    main()
